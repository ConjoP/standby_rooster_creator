import pandas as pd
from datetime import datetime, timedelta
import calendar
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Prompt user to enter people and their phone numbers (supports any number of people)
people = []
print("Enter names and phone numbers for the 'telefoonnummers' sheet.")
print("Type 'done' when finished.")
while True:
    name = input("Enter name (or 'done' to finish): ").strip()
    if name.lower() == "done":
        break
    if not name:
        print("Name cannot be empty. Please try again.")
        continue
    phone = input(f"Enter phone number for {name}: ").strip()
    people.append({"Name": name, "Phone": phone})

# Ask user for the year
while True:
    year_input = input("Enter the year for the template (e.g., 2025): ").strip()
    if year_input.isdigit() and len(year_input) == 4:
        year = int(year_input)
        break
    print("Invalid year. Please enter a 4-digit year.")


# Generate months with correct number of days and start day names
months = []
for month_num in range(1, 13):
    month_name = calendar.month_name[month_num]
    days_in_month = calendar.monthrange(year, month_num)[1]
    start_day_idx = calendar.monthrange(year, month_num)[0]  # 0=Monday
    start_day_name = calendar.day_name[start_day_idx]
    months.append({
        "name": month_name,
        "days": days_in_month,
        "start_day": start_day_name
    })

# Abbreviated day names (Mo, Tu, We, Th, Fr, Sa, Su)
day_abbr = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]

def generate_month_sheet(month_name, days, start_day):
    # Generate day numbers (1, 2, ..., days)
    day_numbers = [str(i) for i in range(1, days + 1)]
    
    # Generate day abbreviations starting from the specified start_day
    start_idx = day_abbr.index(start_day[:2])
    days_of_week = []
    for i in range(days):
        days_of_week.append(day_abbr[(start_idx + i) % 7])
    
    # Create DataFrame structure
    df = pd.DataFrame(
        [
            [f"2025-{month_name}-01 00:00:00"] + day_numbers,
            [""] + days_of_week
        ]
    )
    
    # Add empty rows for each person
    for person in people:
        if person["Name"] == "Back up 1" and month_name not in ["June"]:
            continue  # Skip "Back up 1" except for June
        empty_row = [person["Name"]] + [""] * days
        df.loc[len(df)] = empty_row
    
    return df

# Create Excel writer
with pd.ExcelWriter("Updated_Consignatierooster.xlsx", engine="openpyxl") as writer:
    # Generate all monthly sheets
    for month in months:
        sheet_name = month["name"]
        df = generate_month_sheet(
            sheet_name, month["days"], month["start_day"]
        )
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    
    # Add "telefoonnummers" sheet
    phone_df = pd.DataFrame(
        [
            [person["Name"], "", person["Phone"]]
            for person in people
        ],
        columns=["Name", "", "Phone"]
    )
    phone_df.to_excel(
        writer, sheet_name="telefoonnummers", index=False
    )

# Apply formatting with openpyxl
wb = openpyxl.load_workbook("Updated_Consignatierooster.xlsx")
header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
day_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
name_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
header_font = Font(bold=True, color="000000")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal="center", vertical="center")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    max_col = ws.max_column
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = center_align
    # Color and bold the first row (header)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    # Color the second row (day abbreviations)
    if max_row > 1:
        for cell in ws[2]:
            cell.fill = day_fill
            cell.font = header_font
    # Color the first column (names)
    for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=1):
        for cell in row:
            cell.fill = name_fill
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 30))

wb.save("Updated_Consignatierooster.xlsx")

print("Excel file generated successfully with improved layout!")